import pandas as pd
import os
import sys
import re
import importlib
import zipfile
import datetime
import openpyxl
import shutil
import subprocess
import logging
from pandas import DataFrame, Series
from openpyxl.utils.dataframe import dataframe_to_rows

# pandas options
pd.set_option('display.max_columns', None)  # Shows all columns in DataFrames. See http://pandas.pydata.org/pandas-docs/stable/options.html
pd.set_option('display.max_rows', None) # Shows all rows in DataFrames.
pd.set_option('display.width', 5000)
pd.set_option('display.multi_sparse', False)  #  Display every cell (for multi-level index).
pd.set_option('display.max_colwidth', -1)  # Display full contents of each column.

# Create the logging instance for logging to file only
logger = logging.getLogger('orca_loader')
# Clear existing handlers, else will have duplicate logging messages.
if logger.hasHandlers():
    logger.handlers.clear()
# Create the handler for the main logger
file_logger = logging.FileHandler('orca_loader.log')

str_format = '[%(asctime)s] - [%(levelname)s] - %(message)s'
file_logger_format = logging.Formatter(str_format)
file_logger.setFormatter(file_logger_format)
# Add the handler to the base logger
logger.addHandler(file_logger)
# By default, logging will start at 'warning' unless we tell it otherwise.
logger.setLevel(logging.INFO)


# Copy files from remote source folder to local target folder.
str_source_folder = '//10.0.2.251/FESFTP/feh'  # The remote folder.
str_target_folder = 'C:/AA/python/orca_loader/files/temp'  # The local target folder.

# Pre-requisite: Ensure that there are EXACTLY 5 files in the remote source folder.
l_src_files = os.listdir(str_source_folder)
l_src_files = [f for f in l_src_files if f.endswith('xlsx') or f.startswith('XLSX')]  # Interested only in XLSX files.
if len(l_src_files) != 5:
    logger.error('[PROGRAM TERMINATED] There should be -EXACTLY- 5 .xlsx files in {}'.format(str_source_folder))
    sys.exit()

#print('[COPYING] From {} to {}'.format(str_source_folder, str_target_folder))
logger.info('[COPYING] From {} to {}'.format(str_source_folder, str_target_folder))
for str_fn_src in l_src_files:
    str_fn_full_src = os.path.join(str_source_folder, str_fn_src)
    if (os.path.isfile(str_fn_full_src)):
        if str_fn_full_src.endswith(".xlsx"):  # Take only the XLSX files.
            shutil.copy(src=str_fn_full_src, dst=os.path.join(str_target_folder, str_fn_src))
            
# Global variable definitions #
str_dir = 'C:/AA/python/orca_loader/files/temp'  # Directory containing all the source ZIP files.
str_date = str.upper(datetime.datetime.today().strftime('%d%b%Y'))  # Today's date in prescribed DDMMMYYYY format.

str_cwd = os.getcwd()   # Store existing cwd.
os.chdir(str_dir)  # Change the current working directory.

# Rename all .XLSX files #
#print('[MESSAGE] Renaming files')
logger.info('Renaming files')
# Craft old and new file names for file renaming purposes
l_dir_files = os.listdir(str_dir)
# Filename: CAG (Allotment Groups)
str_fn_old_cag = os.path.join(str_dir, [s for s in l_dir_files if "CAG.xlsx" in s][0])
str_fn_new_cag = os.path.join(str_dir, 'OP_FWD_CAG_' + str_date + '_AM.xlsx')
# Filename: Opera Cancellation
str_fn_old_cancellations = os.path.join(str_dir, [s for s in l_dir_files if "Cancellations.xlsx" in s][0])
str_fn_new_cancellations = os.path.join(str_dir, 'OP_FWD_CAN_' + str_date + '_AM.xlsx')
# Filename: Opera History
str_fn_old_history = os.path.join(str_dir, [s for s in l_dir_files if "History.xlsx" in s][0])
str_fn_new_history = os.path.join(str_dir, 'OP_RES_HIS_' + str_date + '.xlsx')
# Filename: Opera 60 days (Opera 60 days OTB)
str_fn_old_60d = os.path.join(str_dir, [s for s in l_dir_files if "60 days.xlsx" in s][0])
str_fn_new_60d = os.path.join(str_dir, 'OP_RES_' + str_date + '_AM_60.xlsx')
# Filename: Opera 61 days (Opera 61 days onwards OTB)
str_fn_old_61d = os.path.join(str_dir, [s for s in l_dir_files if "61 days.xlsx" in s][0])
str_fn_new_61d = os.path.join(str_dir, 'OP_RES_' + str_date + '_AM_61.xlsx')
# Filename: EzRMS blank template
str_fn_ezrms = os.path.join(str_dir, 'EzRMS_' + str_date + '.xlsx')

os.rename(src=str_fn_old_cag, dst=str_fn_new_cag)
os.rename(src=str_fn_old_cancellations, dst=str_fn_new_cancellations)
os.rename(src=str_fn_old_history, dst=str_fn_new_history)
os.rename(src=str_fn_old_60d, dst=str_fn_new_60d)
os.rename(src=str_fn_old_61d, dst=str_fn_new_61d)

# Rename column name in Opera History and Opera 61 days files #
#print('[MESSAGE] Renaming column names')
logger.info('Renaming column names')
# Opera History
wb = openpyxl.load_workbook(str_fn_new_history)
ws = wb['Guest profile']   # Zoom in to this Worksheet.
for col in range(1, 10):  # Considering only the first row, because the header names are here.
    if ws.cell(row=1, column=col).value is not None:
        ws.cell(row=1, column=col).value = ws.cell(row=1, column=col).value.replace('Special Requests (Codes)', 'Special Requests (Descriptions)')
wb.save(str_fn_new_history)

# Opera 61 days file
wb = openpyxl.load_workbook(str_fn_new_61d)
ws = wb['Guest profile']   # Zoom in to this Worksheet.
for col in range(1, 10):  # Considering only the first row, because the header names are here.
    if ws.cell(row=1, column=col).value is not None:
        ws.cell(row=1, column=col).value = ws.cell(row=1, column=col).value.replace('Special Requests (Codes)', 'Special Requests (Descriptions)')
wb.save(str_fn_new_61d)

# Generate EzRMS blank data template Excel file #
# To run the 2 EzRMS reports -> Export to TSV -> paste into here.
#print('[MESSAGE] Generating EzRMS blank data template file')
logger.info('Generating EzRMS blank data template file')
# Payload to go into each Worksheet. We want to generate the Worksheet headers for convenience.
l_forecast_number = ['Date','DOW','ACH_OCC','CVH_OCC','GLH_OCC','OHS_OCC','OPH_OCC','TES_OCC','TQH_OCC','EVH_OCC','RHS_OCC','OHD_OCC','OKL_OCC','All_OCC','ACH_PFO','CVH_PFO','GLH_PFO','OHS_PFO','OPH_PFO','TES_PFO','TQH_PFO','EVH_PFO','RHS_PFO','OHD_PFO','OKL_PFO','All_PFO']
l_otb = ['Date','DOW','Booking Category','ACH_TY','CVH_TY','GLH_TY','OHS_TY','OPH_TY','TES_TY','TQH_TY','EVH_TY','RHS_TY','OHD_TY','OKL_TY','ACH_LY','CVH_LY','GLH_LY','OHS_LY','OPH_LY','TES_LY','TQH_LY','EVH_LY','RHS_LY','OHD_LY','OKL_LY']
df_forecast_number = DataFrame(columns=l_forecast_number)
df_otb = DataFrame(columns=l_otb)

wb = openpyxl.Workbook()
# Worksheet: 'Forecast Number'
ws = wb.create_sheet('Forecast Number')
for row in dataframe_to_rows(df_forecast_number, index=False, header=True):
    ws.append(row)
for cell in ws[1]:
    cell.style = 'Pandas'
# Worksheet: 'OTB '
ws = wb.create_sheet('OTB ')  # Note the trailing space, because the SAS program is checking for this string!
for row in dataframe_to_rows(df_otb, index=False, header=True):
    ws.append(row)
for cell in ws[1]:
    cell.style = 'Pandas'
wb.save(str_fn_ezrms)

#print('[MESSAGE] ORCA loader pre-processing completed')
logger.info('ORCA loader pre-processing completed')
os.chdir(str_cwd)


# Copy the files to the respective places #
#print('[MESSAGE] Copying files to INTERFACE folders')
logger.info('Copying files to INTERFACE folders')
# shutil.copy(str_fn_ezrms, 'C:\FEOSASBI\INTERFACES\EzRMS')  # Require human intervention to copy-paste.
shutil.copy(str_fn_new_cag, 'C:\FEOSASBI\INTERFACES\OPERA')
shutil.copy(str_fn_new_cancellations, 'C:\FEOSASBI\INTERFACES\OPERA')
shutil.copy(str_fn_new_history, 'C:\FEOSASBI\INTERFACES\OPERA')
shutil.copy(str_fn_new_60d, 'C:\FEOSASBI\INTERFACES\OPERA')
shutil.copy(str_fn_new_61d, 'C:\FEOSASBI\INTERFACES\OPERA')

# Call the external batch file. Can alternatively also call the BAT file created by Emerio, if desired.
# 10 Aug: Shuqi said she'll trigger the SAS job manually first, hence comment out the below.
# p = subprocess.Popen('C:/AA/python/orca_loader/batch/sas_load_op_am.bat', shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
